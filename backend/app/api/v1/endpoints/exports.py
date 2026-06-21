"""
Exports comptables — Journal de ventes OHADA.

GET /exports/journal-ventes?start=YYYY-MM-DD&end=YYYY-MM-DD&format=csv|xlsx
"""
from __future__ import annotations

import csv
import io
from datetime import date
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.v1.deps import CurrentUser
from app.core.database import get_db_session
from app.infrastructure.database.models import (
    ClientModel,
    OrderModel,
    PaymentTransactionModel,
)

router = APIRouter()

_VALID_FORMATS = {"csv", "xlsx"}

# Colonnes OHADA
_HEADERS = [
    "Date",
    "N° Commande",
    "Client",
    "Description",
    "Montant HT (XAF)",
    "TVA",
    "Montant TTC",
    "Statut paiement",
    "Méthode paiement",
]


def _build_rows(transactions: list[PaymentTransactionModel]) -> list[list]:
    rows: list[list] = []
    for txn in transactions:
        order = txn.order
        client_name = ""
        if order and order.client:
            client = order.client
            client_name = client.company_name or client.full_name

        if order:
            descriptions = ", ".join(item.description for item in order.items) if order.items else "—"
            subtotal = order.subtotal_cents
            tax = order.tax_cents
            total = order.total_cents
            payment_status = order.payment_status
        else:
            descriptions = "—"
            subtotal = txn.amount_cents
            tax = 0
            total = txn.amount_cents
            payment_status = txn.status

        rows.append([
            txn.transaction_date.strftime("%d/%m/%Y"),
            order.order_number if order else "—",
            client_name,
            descriptions,
            subtotal,
            tax,
            total,
            payment_status,
            txn.method,
        ])
    return rows


async def _fetch_transactions(
    db: AsyncSession,
    organization_id,
    start: date,
    end: date,
) -> list[PaymentTransactionModel]:
    q = (
        select(PaymentTransactionModel)
        .where(PaymentTransactionModel.organization_id == organization_id)
        .where(PaymentTransactionModel.transaction_date >= start)
        .where(PaymentTransactionModel.transaction_date <= end)
        .options(
            selectinload(PaymentTransactionModel.order).selectinload(OrderModel.items),
            selectinload(PaymentTransactionModel.order).selectinload(OrderModel.client),
        )
        .order_by(PaymentTransactionModel.transaction_date.asc())
    )
    result = await db.execute(q)
    return list(result.scalars().all())


@router.get("/journal-ventes", summary="Journal de ventes OHADA (CSV ou XLSX)")
async def export_journal_ventes(
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db_session)],
    start: date = Query(..., description="Date de début (YYYY-MM-DD)"),
    end: date = Query(..., description="Date de fin (YYYY-MM-DD)"),
    format: str = Query("csv", description="Format : csv ou xlsx"),
):
    if format not in _VALID_FORMATS:
        raise HTTPException(status_code=400, detail=f"Format invalide. Valeurs acceptées : {', '.join(_VALID_FORMATS)}")
    if end < start:
        raise HTTPException(status_code=400, detail="La date de fin doit être postérieure à la date de début.")
    if current_user.organization_id is None:
        raise HTTPException(status_code=403, detail="Cet endpoint requiert un compte rattaché à une organisation.")

    transactions = await _fetch_transactions(db, current_user.organization_id, start, end)
    rows = _build_rows(transactions)

    filename_base = f"journal-ventes-{start}-{end}"

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow(_HEADERS)
        writer.writerows(rows)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_base}.csv"',
            },
        )

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé — contactez l'administrateur.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal de ventes"

    # En-têtes stylisés
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    # Ajustement largeur colonnes
    col_widths = [14, 18, 30, 40, 18, 12, 14, 18, 18]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf_bytes = io.BytesIO()
    wb.save(buf_bytes)
    buf_bytes.seek(0)

    return StreamingResponse(
        iter([buf_bytes.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename_base}.xlsx"',
        },
    )
